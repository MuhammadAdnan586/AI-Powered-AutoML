"""
Excel Summary Exporter
Generates multi-sheet Excel reports: summary, quality, benchmark, features.
"""
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


REPORTS_DIR = Path("artifacts/reports")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


class ExcelExporter:
    """
    Exports AutoML results to a professional multi-sheet Excel workbook.
    """

    # Color scheme
    HEADER_COLOR = "1A237E"   # Dark blue
    ACCENT_COLOR = "283593"   # Medium blue
    SUCCESS_COLOR = "1B5E20"  # Dark green
    WARNING_COLOR = "E65100"  # Orange
    LIGHT_BG = "E8EAF6"       # Light blue-grey
    WHITE = "FFFFFF"

    def __init__(self, output_path: Optional[str] = None):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_path = output_path or str(REPORTS_DIR / f"automl_summary_{ts}.xlsx")
        self.writer = pd.ExcelWriter(self.output_path, engine="openpyxl")

    def export(
        self,
        dataset_info: Dict[str, Any],
        quality_report: Dict[str, Any],
        benchmark_results: Dict[str, Any],
        explanation: Optional[Dict[str, Any]] = None
    ) -> str:
        """Export all data to Excel and return the file path."""

        self._write_summary_sheet(dataset_info, quality_report, benchmark_results)
        self._write_quality_sheet(quality_report)
        self._write_benchmark_sheet(benchmark_results)

        if explanation and explanation.get("feature_importance"):
            self._write_features_sheet(explanation["feature_importance"])

        self.writer.close()
        self._apply_styles()
        return self.output_path

    def _write_summary_sheet(self, dataset_info, quality_report, benchmark_results):
        """Executive summary sheet."""
        summary_data = {
            "Metric": [
                "Dataset Name", "Total Rows", "Total Columns",
                "Task Type", "Target Column",
                "Quality Score", "Quality Grade",
                "Best Model", "Best Model Score",
                "Total Models Trained", "Report Generated"
            ],
            "Value": [
                dataset_info.get("name", "-"),
                dataset_info.get("rows", "-"),
                dataset_info.get("columns", "-"),
                dataset_info.get("task_type", "-"),
                dataset_info.get("target_column", "-"),
                f"{quality_report.get('quality_score', 0)}/100",
                quality_report.get("quality_grade", "-"),
                benchmark_results.get("best_model", {}).get("name", "-"),
                f"{benchmark_results.get('best_model', {}).get('score', 0):.4f}",
                len(benchmark_results.get("models", [])),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }
        pd.DataFrame(summary_data).to_excel(self.writer, sheet_name="Summary", index=False)

    def _write_quality_sheet(self, quality_report):
        """Data quality details sheet."""
        rows = []

        # Missing values
        for col, info in quality_report.get("missing_values", {}).get("columns", {}).items():
            rows.append({
                "Column": col,
                "Issue": "Missing Values",
                "Count": info["missing_count"],
                "Percentage": f"{info['missing_pct']:.2f}%"
            })

        # Outliers
        for col, info in quality_report.get("outliers", {}).items():
            rows.append({
                "Column": col,
                "Issue": "Outliers",
                "Count": info["outlier_count"],
                "Percentage": f"{info['outlier_pct']:.2f}%"
            })

        if rows:
            pd.DataFrame(rows).to_excel(self.writer, sheet_name="Data Quality", index=False)
        else:
            pd.DataFrame({"Status": ["✅ No quality issues found"]}).to_excel(
                self.writer, sheet_name="Data Quality", index=False
            )

        # Recommendations
        recs = quality_report.get("recommendations", [])
        if recs:
            rec_df = pd.DataFrame({"Recommendations": recs})
            rec_df.to_excel(self.writer, sheet_name="Recommendations", index=False)

    def _write_benchmark_sheet(self, benchmark_results):
        """Model benchmark leaderboard sheet."""
        models = benchmark_results.get("models", [])
        if not models:
            return

        rows = []
        for i, m in enumerate(models, 1):
            rows.append({
                "Rank": i,
                "Model": m.get("name", "-"),
                "Score": round(m.get("score", 0), 4),
                "Training Time (s)": round(m.get("training_time", 0), 2),
                "Status": "🏆 Best" if i == 1 else ""
            })

        df = pd.DataFrame(rows)
        df.to_excel(self.writer, sheet_name="Benchmark", index=False)

    def _write_features_sheet(self, feature_importance: list):
        """Feature importance sheet."""
        df = pd.DataFrame(feature_importance)
        df.index = range(1, len(df) + 1)
        df.index.name = "Rank"
        df.to_excel(self.writer, sheet_name="Feature Importance")

    def _apply_styles(self):
        """Apply professional styling to the workbook."""
        wb = load_workbook(self.output_path)

        header_font = Font(name="Calibri", bold=True, color=self.WHITE, size=11)
        header_fill = PatternFill("solid", fgColor=self.HEADER_COLOR)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="BDBDBD"),
            right=Side(style="thin", color="BDBDBD"),
            top=Side(style="thin", color="BDBDBD"),
            bottom=Side(style="thin", color="BDBDBD")
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Style data rows with alternating colors
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                row_color = self.LIGHT_BG if row_idx % 2 == 0 else self.WHITE
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=row_color)
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=10)

            # Auto-fit columns
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col), default=0
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

            # Freeze header row
            ws.freeze_panes = "A2"

        wb.save(self.output_path)
